import openpyxl as xl
from openpyxl.chart import BarChart, Reference
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns

def process_workbook(filename):
    # Load the workbook and the specific sheet
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    # Load data into a pandas DataFrame
    data = pd.read_excel(filename, sheet_name='Sheet1')

    # Data cleaning: Replace non-numeric data in the price column (assuming it's in column 3)
    data['Corrected Price'] = pd.to_numeric(data.iloc[:, 2], errors='coerce') * 0.9
    data.dropna(subset=['Corrected Price'], inplace=True)

    # Save cleaned data back to the Excel sheet
    for index, row in data.iterrows():
        sheet.cell(row=index + 2, column=4).value = row['Corrected Price']

    # Add a Bar Chart using openpyxl
    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')
    wb.save(filename)
    print(f"Workbook {filename} saved successfully.")
    return data

# User input for filename
filename = input("Enter the file name: ")

# Process the workbook and clean the data
data = process_workbook(filename)

# Save cleaned data to a new Excel file for further analysis
cleaned_filename = "cleaned_" + filename
data.to_excel(cleaned_filename, index=False)
print(f"Cleaned data saved to {cleaned_filename}")

# Visualization using Pandas, Matplotlib, and Seaborn

# Load the cleaned data
data = pd.read_excel(cleaned_filename)

# Display the first few rows of the DataFrame to ensure data is loaded correctly
print(data.head())

# 1. Histogram of Corrected Prices
plt.figure(figsize=(10, 6))
sns.histplot(data['Corrected Price'], kde=True, color='blue')
plt.title('Distribution of Corrected Prices')
plt.xlabel('Corrected Price')
plt.ylabel('Frequency')
plt.show()

# 2. Scatter Plot of Corrected Prices
plt.figure(figsize=(10, 6))
sns.scatterplot(x=data.index, y='Corrected Price', data=data, color='green')
plt.title('Corrected Price over Rows')
plt.xlabel('Row Number')
plt.ylabel('Corrected Price')
plt.show()

# 3. Bar Plot of Corrected Prices
plt.figure(figsize=(10, 6))
sns.barplot(x=data.index, y='Corrected Price', data=data, palette='viridis')
plt.title('Bar Plot of Corrected Prices')
plt.xlabel('Row Number')
plt.ylabel('Corrected Price')
plt.show()
